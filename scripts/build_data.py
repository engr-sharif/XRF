#!/usr/bin/env python3
"""
Build site/data/data.json from the master XLSX tracker and the gun CSVs.

Inputs (repo root):
  SBMM_XRF_Sample_Tracker_*.xlsx
  SBM-XRF-<GUN>_Master_*.csv

Output:
  site/data/data.json
  site/data/<csv files>  (copied verbatim for client-side export)
"""

import csv
import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SITE_DATA = ROOT / "site" / "data"

PERIODIC = {
    "H","He","Li","Be","B","C","N","O","F","Ne","Na","Mg","Al","Si","P","S","Cl","Ar",
    "K","Ca","Sc","Ti","V","Cr","Mn","Fe","Co","Ni","Cu","Zn","Ga","Ge","As","Se","Br","Kr",
    "Rb","Sr","Y","Zr","Nb","Mo","Tc","Ru","Rh","Pd","Ag","Cd","In","Sn","Sb","Te","I","Xe",
    "Cs","Ba","La","Ce","Hf","Ta","W","Re","Os","Ir","Pt","Au","Hg","Tl","Pb","Bi","Th","U",
}

# CSV files → (gun serial, file label)
CSV_FILES = [
    ("SBM-XRF-X500456_Master_AllModes.csv",            "X500456", "AllModes"),
    ("SBM-XRF-X500456_Master_SystemChecks.csv",        "X500456", "SystemChecks"),
    ("SBM-XRF-X501203_Master_Mining_Rdg0001-0079.csv", "X501203", "Mining"),
    ("SBM-XRF-X501203_Master_Mining_Rdg0080-0939.csv", "X501203", "Mining"),
    ("SBM-XRF-X501203_Master_Mining_Rdg0940-1868.csv", "X501203", "Mining"),
    ("SBM-XRF-X501203_Master_OtherModes.csv",          "X501203", "OtherModes"),
]


def jsonify(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, float) and (v != v):  # NaN
        return None
    return v


def parse_float(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    t = str(s).strip()
    if not t or t == "<LOD":
        return None
    # Some columns are "value , 2sigma" — main value is first
    if "," in t:
        t = t.split(",")[0].strip()
    try:
        return float(t)
    except ValueError:
        return None


def is_lod(s):
    return isinstance(s, str) and s.strip() == "<LOD"


# ---------------------------------------------------------------------------
# Load gun CSVs into a unified per-reading record keyed by (gun, reading_no)
# ---------------------------------------------------------------------------

def load_csvs():
    """Returns dict[(gun, rdg_no)] = reading record."""
    readings = {}
    all_elements = set()

    for fname, gun, label in CSV_FILES:
        path = ROOT / fname
        if not path.exists():
            print(f"  ! missing {fname}", file=sys.stderr)
            continue

        with path.open(newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            cols = next(reader)
            # Identify element columns: bare symbol (no suffix)
            elem_cols = {}      # col_index -> element symbol
            lod_cols = {}       # col_index -> element symbol (the 2-sigma/LOD col)
            for i, c in enumerate(cols):
                c_strip = c.strip()
                # bare element (e.g., "As", "Pb")
                if c_strip in PERIODIC:
                    elem_cols[i] = c_strip
                    all_elements.add(c_strip)
                elif c_strip.endswith(" 2-Sigma/LOD"):
                    sym = c_strip.split(" ")[0]
                    if sym in PERIODIC:
                        lod_cols[i] = sym
            # Find common metadata columns
            def col(name):
                try:
                    return cols.index(name)
                except ValueError:
                    return None

            c_rdg = col("Reading No")
            c_type = col("Reading Type")
            c_mode = col("Mode")
            c_time = col("Time")
            c_dur = col("Duration")
            c_units = col("Units")
            c_sample = col("Sample")
            c_pf = col("Pass/Fail")
            c_pfc = col("Pass/Fail Criteria")
            c_note = col("Note")
            c_user = col("User")
            c_sigma = col("Sigma Value")
            c_lat = col("Latitude")
            c_lon = col("Longitude")
            c_east = col("Easting")
            c_north = col("Northing")
            c_alt = col("Altitude")

            for row in reader:
                if not row or len(row) < 2 or not row[c_rdg]:
                    continue
                try:
                    rdg_no = int(row[c_rdg])
                except (ValueError, TypeError):
                    continue
                key = (gun, rdg_no)

                elems = {}
                lod = {}
                for ci, sym in elem_cols.items():
                    raw = row[ci] if ci < len(row) else None
                    if raw is None or raw == "":
                        continue
                    if is_lod(raw):
                        lod[sym] = True
                    else:
                        v = parse_float(raw)
                        if v is not None:
                            elems[sym] = v

                # 2-sigma/LOD values
                lod_vals = {}
                for ci, sym in lod_cols.items():
                    raw = row[ci] if ci < len(row) else None
                    v = parse_float(raw)
                    if v is not None:
                        lod_vals[sym] = v

                rec = {
                    "gun": gun,
                    "file": label,
                    "rdgNo": rdg_no,
                    "readingType": row[c_type] if c_type is not None and c_type < len(row) else None,
                    "mode": row[c_mode] if c_mode is not None and c_mode < len(row) else None,
                    "time": row[c_time] if c_time is not None and c_time < len(row) else None,
                    "duration": parse_float(row[c_dur]) if c_dur is not None and c_dur < len(row) else None,
                    "units": row[c_units] if c_units is not None and c_units < len(row) else None,
                    "sampleField": row[c_sample] if c_sample is not None and c_sample < len(row) else None,
                    "passFail": row[c_pf] if c_pf is not None and c_pf < len(row) else None,
                    "passFailCriteria": row[c_pfc] if c_pfc is not None and c_pfc < len(row) else None,
                    "note": row[c_note] if c_note is not None and c_note < len(row) else None,
                    "user": row[c_user] if c_user is not None and c_user < len(row) else None,
                    "sigmaValue": parse_float(row[c_sigma]) if c_sigma is not None and c_sigma < len(row) else None,
                    "lat": parse_float(row[c_lat]) if c_lat is not None and c_lat < len(row) else None,
                    "lon": parse_float(row[c_lon]) if c_lon is not None and c_lon < len(row) else None,
                    "easting": parse_float(row[c_east]) if c_east is not None and c_east < len(row) else None,
                    "northing": parse_float(row[c_north]) if c_north is not None and c_north < len(row) else None,
                    "altitude": parse_float(row[c_alt]) if c_alt is not None and c_alt < len(row) else None,
                    "elements": elems,
                    "lod": lod,
                    "lodValues": lod_vals,
                }
                readings[key] = rec

    return readings, sorted(all_elements)


# ---------------------------------------------------------------------------
# Read tracker workbook
# ---------------------------------------------------------------------------

def header_row(ws, header_row_idx):
    """Get header row, joining multi-row headers if needed."""
    return [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]


def normalize_header(h):
    if h is None:
        return None
    return re.sub(r"\s+", " ", str(h)).strip()


def sheet_rows(ws, header_row_idx, start_row):
    headers = [normalize_header(h) for h in header_row(ws, header_row_idx)]
    out = []
    for r in range(start_row, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            if h:
                rec[h] = jsonify(v)
        out.append(rec)
    return out


def load_tracker():
    xlsx = next(ROOT.glob("SBMM_XRF_Sample_Tracker_*.xlsx"))
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # Boulder Sample Log — header row 4, data starts row 5
    bsl = sheet_rows(wb["Boulder Sample Log"], 4, 5)
    # Drop rows that don't have a sample ID
    bsl = [r for r in bsl if r.get("Sample ID")]

    # XRF Readings Detail — header 4, data row 5
    xrd = sheet_rows(wb["XRF Readings Detail"], 4, 5)
    xrd = [r for r in xrd if r.get("Sample ID") and r.get("XRF Rdg #") is not None]

    # Boulder Selection Log (depth boulders) — header 4
    bsel = sheet_rows(wb["Boulder Selection Log"], 4, 5)
    bsel = [r for r in bsel if r.get("Sample ID")]

    # Depth XRF Detail — header 4
    dxd = sheet_rows(wb["Depth XRF Detail"], 4, 5)
    dxd = [r for r in dxd if r.get("Sample ID") and r.get("XRF Rdg #") is not None]

    # Powder XRF Detail — header 4
    pxd = sheet_rows(wb["Powder XRF Detail"], 4, 5)
    pxd = [r for r in pxd if r.get("XRF Rdg #") is not None]

    # Lab Samples — header 4
    lab = sheet_rows(wb["Lab Samples"], 4, 5)
    lab = [r for r in lab if r.get("Powder Sample ID")]

    # Daily Verification — header 4 (row 5 is certified values, row 6 onward is data)
    ver = sheet_rows(wb["Daily Verification"], 4, 6)
    cert_row_raw = [wb["Daily Verification"].cell(row=5, column=c).value for c in range(1, wb["Daily Verification"].max_column + 1)]
    ver = [r for r in ver if r.get("Date") and r.get("Date") != "CERTIFIED VALUES →"]

    # Daily Field Log — header row 3, data row 4
    field = sheet_rows(wb["Daily Field Log"], 3, 4)
    field = [r for r in field if r.get("Date")]

    # Grid Completion — special: header row 3, area rows 4-10 (7 areas), totals at 11
    gc_ws = wb["Grid Completion"]
    grid_rows = []
    for r in range(4, 11):
        row = [gc_ws.cell(row=r, column=c).value for c in range(1, gc_ws.max_column + 1)]
        if row[0] and row[0] != "TOTAL":
            grid_rows.append(row)

    return {
        "boulders": bsl,
        "readings": xrd,
        "depthBoulders": bsel,
        "depthReadings": dxd,
        "powderReadings": pxd,
        "labSamples": lab,
        "verifications": ver,
        "verificationCert": [jsonify(v) for v in cert_row_raw],
        "fieldLog": field,
        "gridCompletion": grid_rows,
    }


# ---------------------------------------------------------------------------
# Build hierarchical area → grid → boulders structure & join readings
# ---------------------------------------------------------------------------

SAMPLE_ID_RE = re.compile(r"^SBM-([A-Z]+)-G(\d+)-B(\d+)")

# Some sheets (notably Depth XRF Detail) use variant boulder IDs with an
# inserted -HB-/-FB- segment and/or an unpadded grid number. Boulder Sample Log
# canonicalizes on SBM-{AREA}-G{NN}-B{n}, so normalize variants to that form
# before joining.
SAMPLE_ID_NORMALIZE_RE = re.compile(r"^(SBM-[A-Z]+)-G(\d+)(?:-[HF]B)?-B(\d+)$")


def normalize_sample_id(sid):
    if not sid:
        return sid
    s = str(sid).strip()
    m = SAMPLE_ID_NORMALIZE_RE.match(s)
    if m:
        return f"{m.group(1)}-G{int(m.group(2)):02d}-B{m.group(3)}"
    return s


def parse_sample_id(sid):
    if not sid:
        return None
    m = SAMPLE_ID_RE.match(str(sid))
    if not m:
        return None
    return {"area": m.group(1), "grid": int(m.group(2)), "boulder": int(m.group(3))}


def build_areas(tracker):
    # Areas from Grid Completion sheet, in order
    areas = []
    grid_status = {}  # (area, grid) -> "✓" / "½" / None
    for row in tracker["gridCompletion"]:
        area = row[0]
        if not area:
            continue
        grids_status = {}
        for i, status in enumerate(row[1:16]):
            grids_status[i + 1] = status
            grid_status[(area, i + 1)] = status
        areas.append({
            "code": area,
            "gridsStatus": grids_status,
            "done": row[16],
            "notes": row[17],
        })
    return areas, grid_status


def join_readings(tracker, gun_readings):
    """Annotate tracker readings with full element data from gun CSVs."""
    # XRF Readings Detail entries — these are surface readings
    surface = []
    for r in tracker["readings"]:
        gun = r.get("XRF S/N")
        rdg = r.get("XRF Rdg #")
        if gun is None or rdg is None:
            continue
        key = (gun, int(rdg))
        csv_row = gun_readings.get(key)
        rec = dict(r)
        rec["_key"] = f"{gun}:{rdg}"
        rec["_csv"] = csv_row
        surface.append(rec)

    depth = []
    for r in tracker["depthReadings"]:
        gun = r.get("XRF S/N")
        rdg = r.get("XRF Rdg #")
        if gun is None or rdg is None:
            continue
        key = (gun, int(rdg))
        csv_row = gun_readings.get(key)
        rec = dict(r)
        # Normalize Sample ID so depth-XRF variant naming (G6-HB-B1, G06-FB-B2,
        # etc.) joins to the canonical Boulder Sample Log ID (G06-B1, G06-B2).
        original_sid = rec.get("Sample ID")
        normalized = normalize_sample_id(original_sid)
        if normalized != original_sid:
            rec["Sample ID (raw)"] = original_sid
            rec["Sample ID"] = normalized
        rec["_key"] = f"{gun}:{rdg}"
        rec["_csv"] = csv_row
        depth.append(rec)

    powder = []
    # Powder readings need to inherit Sample ID, Parent Boulder, Date from preceding row if blank
    last_sid = last_parent = last_date = None
    for r in tracker["powderReadings"]:
        gun = r.get("XRF S/N")
        rdg = r.get("XRF Rdg #")
        sid = r.get("Sample ID") or last_sid
        parent = r.get("Parent Boulder") or last_parent
        d = r.get("Date") or last_date
        if r.get("Sample ID"):
            last_sid = r["Sample ID"]; last_parent = r.get("Parent Boulder"); last_date = r.get("Date")
        if gun is None or rdg is None:
            continue
        key = (gun, int(rdg))
        csv_row = gun_readings.get(key)
        rec = dict(r)
        rec["Sample ID"] = sid
        rec["Parent Boulder"] = normalize_sample_id(parent) if parent else parent
        if parent and rec["Parent Boulder"] != parent:
            rec["Parent Boulder (raw)"] = parent
        rec["Date"] = d
        rec["_key"] = f"{gun}:{rdg}"
        rec["_csv"] = csv_row
        powder.append(rec)

    return surface, depth, powder


def main():
    print("Loading gun CSVs...")
    gun_readings, all_elements = load_csvs()
    print(f"  {len(gun_readings)} readings across {len({k[0] for k in gun_readings})} guns, {len(all_elements)} elements")

    print("Loading tracker XLSX...")
    tracker = load_tracker()
    print(f"  {len(tracker['boulders'])} boulders, {len(tracker['readings'])} surface rdgs,"
          f" {len(tracker['depthReadings'])} depth rdgs, {len(tracker['powderReadings'])} powder rdgs")

    print("Joining tracker readings to gun CSVs...")
    surface, depth, powder = join_readings(tracker, gun_readings)
    matched_surface = sum(1 for r in surface if r["_csv"])
    matched_depth = sum(1 for r in depth if r["_csv"])
    matched_powder = sum(1 for r in powder if r["_csv"])
    print(f"  surface matched {matched_surface}/{len(surface)}, depth {matched_depth}/{len(depth)}, powder {matched_powder}/{len(powder)}")

    areas, grid_status = build_areas(tracker)

    # Find readings in CSV that are not in tracker (orphans)
    tracker_keys = set()
    for r in surface + depth + powder:
        tracker_keys.add(r["_key"])
    # Verifications keys
    for v in tracker["verifications"]:
        gun = v.get("XRF S/N")
        rdg = v.get("XRF Rdg #")
        if gun and rdg is not None:
            tracker_keys.add(f"{gun}:{int(rdg)}")

    orphans = []
    for (gun, rdg), csv_row in gun_readings.items():
        if f"{gun}:{rdg}" not in tracker_keys:
            orphans.append(csv_row)

    print(f"  orphan readings in CSVs (not in tracker): {len(orphans)}")

    # Build boulder index
    boulders_by_id = {}
    for b in tracker["boulders"]:
        sid = b.get("Sample ID")
        if not sid:
            continue
        boulders_by_id[sid] = b

    # Attach depth-boulder metadata
    for db in tracker["depthBoulders"]:
        sid = db.get("Sample ID")
        if sid in boulders_by_id:
            boulders_by_id[sid]["_depthMeta"] = db

    out = {
        "meta": {
            "builtAt": datetime.utcnow().isoformat() + "Z",
            "guns": sorted({k[0] for k in gun_readings}),
            "elements": all_elements,
            "defaultElements": ["As", "Pb", "Hg", "Sb", "Cu", "Zn", "Fe", "Mn"],
            "areas": [a["code"] for a in areas],
            "counts": {
                "boulders": len(tracker["boulders"]),
                "surfaceReadings": len(surface),
                "depthReadings": len(depth),
                "powderReadings": len(powder),
                "labSamples": len(tracker["labSamples"]),
                "verifications": len(tracker["verifications"]),
                "fieldLogDays": len(tracker["fieldLog"]),
                "csvReadingsTotal": len(gun_readings),
                "orphans": len(orphans),
            },
            "csvFiles": [{"file": f, "gun": g, "label": l} for f, g, l in CSV_FILES],
        },
        "areas": areas,
        "boulders": list(boulders_by_id.values()),
        "surfaceReadings": surface,
        "depthReadings": depth,
        "powderReadings": powder,
        "labSamples": tracker["labSamples"],
        "verifications": tracker["verifications"],
        "verificationCert": tracker["verificationCert"],
        "fieldLog": tracker["fieldLog"],
        "orphans": orphans,
    }

    SITE_DATA.mkdir(parents=True, exist_ok=True)
    out_path = SITE_DATA / "data.json"
    with out_path.open("w") as f:
        json.dump(out, f, default=jsonify, separators=(",", ":"))
    print(f"  wrote {out_path} ({out_path.stat().st_size / 1024:.0f} KB)")

    # Copy CSVs into site/data for client-side export
    for fname, _gun, _label in CSV_FILES:
        src = ROOT / fname
        if src.exists():
            shutil.copy(src, SITE_DATA / fname)
    print(f"  copied {len(CSV_FILES)} CSVs to site/data/")


if __name__ == "__main__":
    main()
